"""
Export TTS comparison data to Excel (.xlsx) or CSV fallback.
Run: python3 export_to_excel.py
"""

import sys

MODELS = [
    {
        "Model": "AWS Nova Sonic v2",
        "HuggingFace Link": "https://aws.amazon.com/ai/generative-ai/nova/sonic/",
        "Size": "Not disclosed",
        "Latency": "Ultra-low (streaming)",
        "Streaming / Full-Duplex": "Yes – full-duplex",
        "Mac Compatible": "Cloud only",
        "Voice Cloning": "Doesn't say",
        "Emotions": "Yes",
        "Watermarking": "Yes",
        "Semantic Extraction": "Yes (claimed, no clear verification)",
        "License": "Proprietary",
        "Language Support": "Multilingual",
        "Notable Features": "Prompt-injectable emotions and filler words e.g. [laugh] That's funny",
        "Personal Audio Notes": "Very natural",
    },
    {
        "Model": "NVIDIA PersonaPlex",
        "HuggingFace Link": "https://huggingface.co/nvidia/personaplex-7b-v1",
        "Size": "7B",
        "Latency": "170ms (turn-taking) / 240ms (interruption)",
        "Streaming / Full-Duplex": "Yes – full-duplex concurrent listen/speak",
        "Mac Compatible": "Linux + NVIDIA GPU only",
        "Voice Cloning": "Yes (voice prompt control, 0.650 SSIM)",
        "Emotions": "No",
        "Watermarking": "Doesn't say",
        "Semantic Extraction": "Yes",
        "License": "NVIDIA Open Model License",
        "Language Support": "English only (24kHz)",
        "Notable Features": "Dual-stream architecture; 90.8% smooth turn-taking; text prompt persona control",
        "Personal Audio Notes": "N/A — not tested locally",
    },
    {
        "Model": "VibeVoice-Realtime-0.5B",
        "HuggingFace Link": "https://huggingface.co/microsoft/VibeVoice-Realtime-0.5B",
        "Size": "0.5B",
        "Latency": "~300ms",
        "Streaming / Full-Duplex": "Yes – streaming text-in / audio-out",
        "Mac Compatible": "Yes",
        "Voice Cloning": "No (single speaker only)",
        "Emotions": "No",
        "Watermarking": "Yes (imperceptible + audible AI disclaimer)",
        "Semantic Extraction": "Doesn't say",
        "License": "MIT",
        "Language Support": "English (primary) + 9 experimental: German, French, Italian, Japanese, Korean, Dutch, Polish, Portuguese, Spanish",
        "Notable Features": "Acoustic tokenizer 3200x downsampling; diffusion-based decoding; up to 10 min continuous; multi-speaker TTS variant supports 90-min (full podcast)",
        "Personal Audio Notes": "Surprisingly natural — almost the same feel as Nova Sonic 2",
    },
    {
        "Model": "Chatterbox",
        "HuggingFace Link": "https://huggingface.co/ResembleAI/chatterbox",
        "Size": "0.5B",
        "Latency": "<200ms",
        "Streaming / Full-Duplex": "Not explicitly stated",
        "Mac Compatible": "Yes",
        "Voice Cloning": "Yes (zero-shot voice conversion)",
        "Emotions": "Yes (emotion exaggeration control)",
        "Watermarking": "Yes (Perth watermarking — survives MP3 compression)",
        "Semantic Extraction": "Doesn't say",
        "License": "MIT",
        "Language Support": "Arabic, Danish, German, Greek, English, Spanish, Finnish, French, Hebrew, Hindi, Italian, Japanese, Korean, Malay, Dutch, Norwegian, Polish, Portuguese, Russian, Swedish, Swahili, Turkish, Chinese (23 languages)",
        "Notable Features": "0.5B Llama backbone; 0.5M hours training data; first open-source TTS with emotion exaggeration; CFG/Pace control",
        "Personal Audio Notes": "Very natural when you tune the correct exaggeration and weights",
    },
    {
        "Model": "CosyVoice 3",
        "HuggingFace Link": "https://huggingface.co/FunAudioLLM/Fun-CosyVoice3-0.5B-2512",
        "Size": "0.5B",
        "Latency": "~150ms",
        "Streaming / Full-Duplex": "Yes – bi-streaming (text-in & audio-out)",
        "Mac Compatible": "Not yet implemented — too tedious",
        "Voice Cloning": "Yes (zero-shot + cross-lingual)",
        "Emotions": "Yes (instruct-based, e.g. [laugh])",
        "Watermarking": "Doesn't say",
        "Semantic Extraction": "Doesn't say",
        "License": "Apache 2.0",
        "Language Support": "Chinese, English, Japanese, Korean, German, Spanish, French, Italian, Russian + 18+ Chinese dialects",
        "Notable Features": "Pronunciation inpainting (Pinyin + CMU); 77.4-78% speaker similarity; 1.68% WER",
        "Personal Audio Notes": "Not yet tested — implementation too tedious",
    },
    {
        "Model": "Dia-2B",
        "HuggingFace Link": "https://huggingface.co/nari-labs/Dia2-1B",
        "Size": "1B or 2B (two variants)",
        "Latency": "Can start with first few words",
        "Streaming / Full-Duplex": "Yes – starts generating without full text",
        "Mac Compatible": "Yes",
        "Voice Cloning": "Yes (conditional via speaker prefixes)",
        "Emotions": "Limited (speaker tags [S1] [S2] only)",
        "Watermarking": "No",
        "Semantic Extraction": "No",
        "License": "Apache 2.0",
        "Language Support": "English only (max 2 min audio)",
        "Notable Features": "Two-persona dialogue model; word-level timestamps; Kyutai Mimi codec",
        "Personal Audio Notes": "Multiple persona dialogue sounds quite natural",
    },
    {
        "Model": "DiVA",
        "HuggingFace Link": "https://huggingface.co/WillHeld/DiVA-llama-3-v0-8b",
        "Size": "8B",
        "Latency": "Not specified",
        "Streaming / Full-Duplex": "No",
        "Mac Compatible": "Yes — very difficult, took too long",
        "Voice Cloning": "No",
        "Emotions": "No",
        "Watermarking": "No",
        "Semantic Extraction": "No",
        "License": "Apache 2.0",
        "Language Support": "Multiple (CommonVoice dataset, mostly English in practice)",
        "Notable Features": "End-to-end voice assistant (speech + text); distillation-based; multimodal; based on Llama-3.1-8B",
        "Personal Audio Notes": "Ugly, robotic",
    },
    {
        "Model": "fishspeech S1 mini",
        "HuggingFace Link": "https://huggingface.co/fishaudio/s1-mini",
        "Size": "0.5B (mini) / 4B (full)",
        "Latency": "Very slow on Mac (designed for CUDA)",
        "Streaming / Full-Duplex": "No",
        "Mac Compatible": "Yes — extremely annoying to set up",
        "Voice Cloning": "Yes",
        "Emotions": "Yes (49 emotion + tone markers)",
        "Watermarking": "No",
        "Semantic Extraction": "No",
        "License": "CC-BY-NC-SA-4.0 (NON-COMMERCIAL ONLY)",
        "Language Support": "English, Chinese, Japanese, German, French, Spanish, Korean, Arabic, Russian, Dutch, Italian, Polish, Portuguese (13 languages)",
        "Notable Features": "2M+ hours training; online RLHF; emotion tags: (laughing)(sobbing)(whispering)(angry)(sad)(excited)(surprised) and 40+ more",
        "Personal Audio Notes": "Sounds funny",
    },
    {
        "Model": "Kimi-Audio-7B-Instruct",
        "HuggingFace Link": "https://huggingface.co/moonshotai/Kimi-Audio-7B-Instruct",
        "Size": "7B",
        "Latency": "Not specified",
        "Streaming / Full-Duplex": "Yes (chunk-wise streaming detokenizer)",
        "Mac Compatible": "No – CUDA only",
        "Voice Cloning": "No",
        "Emotions": "Yes (SER — speech emotion recognition)",
        "Watermarking": "No",
        "Semantic Extraction": "Yes (ASR, AQA, AAC, SER, SEC/ASC)",
        "License": "MIT / Apache 2.0",
        "Language Support": "English, Chinese (13M+ hours training data)",
        "Notable Features": "ASR, audio QA, audio captioning, SER, sound event/scene classification; hybrid acoustic + semantic tokens",
        "Personal Audio Notes": "Not tested",
    },
    {
        "Model": "Kokoro-82M TTS",
        "HuggingFace Link": "https://huggingface.co/hexgrad/Kokoro-82M",
        "Size": "82M",
        "Latency": "Blazing fast locally (no official claim)",
        "Streaming / Full-Duplex": "No",
        "Mac Compatible": "Yes",
        "Voice Cloning": "No (explicitly excluded)",
        "Emotions": "No",
        "Watermarking": "No",
        "Semantic Extraction": "No",
        "License": "Apache 2.0",
        "Language Support": "American English, British English, Japanese, Mandarin Chinese, Spanish, French, Hindi, Italian, Brazilian Portuguese — 54 voices total",
        "Notable Features": "Extremely small (82M) but natural-sounding; StyleTTS 2 + ISTFTNet vocoder; trained on ~$1,000 compute",
        "Personal Audio Notes": "Very natural — surprising",
    },
    {
        "Model": "Magpie TTS 357M",
        "HuggingFace Link": "https://huggingface.co/nvidia/magpie_tts_multilingual_357m",
        "Size": "357M",
        "Latency": "Extremely fast (no official claim)",
        "Streaming / Full-Duplex": "Yes (streaming voice agent, 20s chunks)",
        "Mac Compatible": "Yes",
        "Voice Cloning": "No (zero-shot explicitly removed)",
        "Emotions": "No (enterprise NIM only, not in open model)",
        "Watermarking": "No",
        "Semantic Extraction": "No",
        "License": "NVIDIA Open Model License",
        "Language Support": "English, Spanish, German, French, Vietnamese, Italian, Mandarin Chinese, Hindi, Japanese (9 languages)",
        "Notable Features": "5 fixed speakers; multi-codebook prediction; local transformer refinement; trained on 50K hours; built-in text normalization",
        "Personal Audio Notes": "Very natural — surprising",
    },
    {
        "Model": "Ming-Omni-TTS-0.5B",
        "HuggingFace Link": "https://huggingface.co/inclusionAI/Ming-omni-tts-0.5B",
        "Size": "0.5B",
        "Latency": "3.1Hz (Patch-by-Patch compression)",
        "Streaming / Full-Duplex": "Not stated",
        "Mac Compatible": "Yes",
        "Voice Cloning": "Yes (zero-shot)",
        "Emotions": "Yes (fine-grained, 76.7% accuracy)",
        "Watermarking": "No",
        "Semantic Extraction": "No",
        "License": "Apache 2.0",
        "Language Support": "Chinese (primary), English, Cantonese",
        "Notable Features": "Unified speech + ambient sound + music generation; handles math/chemical equations; vocal controls: rate, pitch, volume, emotion, dialect; 0.83% WER (Chinese)",
        "Personal Audio Notes": "Ugly",
    },
    {
        "Model": "CSM-1B",
        "HuggingFace Link": "https://huggingface.co/sesame/csm-1b",
        "Size": "1B",
        "Latency": "Slow",
        "Streaming / Full-Duplex": "Not stated",
        "Mac Compatible": "Yes",
        "Voice Cloning": "Limited (base model — requires fine-tuning for specific voices)",
        "Emotions": "No",
        "Watermarking": "No",
        "Semantic Extraction": "No",
        "License": "Apache 2.0",
        "Language Support": "English (primary; limited non-English due to data contamination)",
        "Notable Features": "Context-aware generation; multi-speaker dialogue; batch inference; CUDA graph compilation; native Transformers integration",
        "Personal Audio Notes": "Sounds great",
    },
]

COLUMNS = [
    "Model",
    "Size",
    "Latency",
    "Streaming / Full-Duplex",
    "Mac Compatible",
    "Voice Cloning",
    "Emotions",
    "Watermarking",
    "Semantic Extraction",
    "License",
    "Language Support",
    "Notable Features",
    "Personal Audio Notes",
    "HuggingFace Link",
]


def export_xlsx(output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "TTS Comparison"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Write data rows
    for row_idx, model in enumerate(MODELS, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = model.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # Auto-size columns (approximate)
    col_widths = {
        "Model": 28,
        "Size": 14,
        "Latency": 30,
        "Streaming / Full-Duplex": 30,
        "Mac Compatible": 28,
        "Voice Cloning": 32,
        "Emotions": 30,
        "Watermarking": 36,
        "Semantic Extraction": 28,
        "License": 28,
        "Language Support": 55,
        "Notable Features": 60,
        "Personal Audio Notes": 40,
        "HuggingFace Link": 55,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Excel file saved: {output_path}")


def export_csv(output_path: str) -> None:
    import csv

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(MODELS)
    print(f"CSV file saved: {output_path}")


if __name__ == "__main__":
    import os

    base = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(base, "TTS_Comparison.xlsx")
    csv_path = os.path.join(base, "TTS_Comparison.csv")

    try:
        export_xlsx(xlsx_path)
    except ImportError:
        print("openpyxl not found — falling back to CSV.")
        print("Install with: pip3 install openpyxl")
        export_csv(csv_path)
